# Golden gate of level 1 plasmids using the MoClo YTK design



# See LICENSE for full GpLv2 license.

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License or
# (at your option) any later version.

# This program is distributed in the hope that it will be useful;,
# but WITHOUT ANY WARRENTY; without even the implied warranty of
# MERCHANTABILITY of FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License along
# with this program; if not, write to the Free Software Foundation, Inc.,
# 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.

# ===========================================================================

# src/pyeast/core/gg_lvl1.py
"""
Level 1 golden gate assemblies using the yeast Moclo standard from Lee et al 2015
"""

# ===========================================================================



import csv
import logging
from pathlib import Path
from typing import Optional

import dnacauldron as dc
import openpyxl
import pandas as pd
from Bio import BiopythonDeprecationWarning, SeqIO
from Bio.SeqRecord import SeqRecord

from pyeast.utils.path_utils import get_private_equivalent, get_templates_path
from pyeast.utils.sequence_utils import get_templates, load_sequences

logger = logging.getLogger(__name__)


class ggDesigner:
    """A class for designing golden gate assemblies.

    This class handles the design assemblies. Uses the DNAcauldron package from the Edinborough bifoundry.
    """

    def __init__(self,
                 template_folder: Optional[Path] = None,
                 instruments: list = ['Janus', 'epMotion', 'Hamilton', 'Human'],
                 is_library: bool = False
                 ):
        """Inintialise a new gg_lvl1 designer.

        Args:
            template_folder:
                Path to a fold where the file catalogging plasmid positions in 96 well plates,
                 TemPlates.xlsx, is stored. Defaults to data/templates.

            Instruments:
                List of available instruments, each requires implementation so shouldn't be input by the user.

            is_library:
                Bool that flags if the assembly is to be multiplexed in separate wells, or pooled to create a library.
                Defaults to False to return separate assemblies.
            """
        # File paths                     # These need to be user determined.
        self.gg_plasmids = None
        self.template_folder = template_folder if template_folder is not None else get_templates_path()

        # state storage
        self.instruments = instruments
        self.multiplex = False
        self.is_library = is_library
        self.available_sequences = {}     # All loaded sequences
        self.assembly_sequences = []      # Sequences in assembly order may be multiple assemblies for multiplex work (list of lists)
        self.assemblies_names = None      # Part names stored as strings
        self.template_dict = {}           # Template information
        self.repository = None            # dnacauldron object constructed from all lvl 0 parts plasmids for gg assembly
        self.assembly = None              # Assembly object from dnacauldron
        self.assembly_sim = None          # Simulation object from dnacauldron
        self.final_assembly = None        # Final assembled sequence(s)


    def load_and_get_sequences(self, directory: Path) -> None:
        """"Loads seqeunces from a directory and store them for assembly.

        Args:
            directory: Path to a directory containing fasta files containing DNA components
            as fasta files
            """
        self.available_sequences = load_sequences(directory)
        self.gg_plasmids = Path(f"{directory}/plasmids")
        return self.available_sequences

    def get_plasmid_names(self) -> list[str]:
        """Map sequences to plasmids containing those sequences.

        Searches for plasmids in both public and private plasmid directories.
        Sets self.plasmid_names for dnacauldron assembly.

        Returns:
            List[str] of plasmid filenames (not paths) for dnacauldron assembly
        Raises:
            ValueError: If no assemblies have been selected
            FileNotFoundError: If no plasmid folder exists
            RuntimeError: If parts cannot be mapped to plasmids
        """
        if not hasattr(self, 'assemblies_names') or not self.assemblies_names:
            raise ValueError("No assemblies selected. Please run get_assembly_order first")

        if not hasattr(self, 'assemblies_names') or not self.assemblies_names:
            raise ValueError("No assembly sequence available. Please run get_seq_records first.")

        # Check both public and private plasmid directories
        public_plasmids = self.gg_plasmids

        # Construct private plasmids path (if path is within data directory)
        try:
            private_plasmids = get_private_equivalent(public_plasmids)
        except ValueError:
            # Path is not within data directory, so no private equivalent exists
            private_plasmids = None

        # Collect plasmid files from both locations
        plasmid_files = []
        found_any_dir = False

        if public_plasmids.exists():
            found_any_dir = True
            plasmid_files.extend(list(public_plasmids.glob('*.gb')))
            plasmid_files.extend(list(public_plasmids.glob('*.gbk')))

        if private_plasmids and private_plasmids.exists():
            found_any_dir = True
            plasmid_files.extend(list(private_plasmids.glob('*.gb')))
            plasmid_files.extend(list(private_plasmids.glob('*.gbk')))

        if not found_any_dir:
            if private_plasmids:
                raise FileNotFoundError(f"Plasmid folder not found in {public_plasmids} or {private_plasmids}")
            else:
                raise FileNotFoundError(f"Plasmid folder not found in {public_plasmids}")

        # Collect all the unique sequence names from all assemblies
        all_part_names = set()
        for assembly in self.assemblies_names:
            all_part_names.update(assembly)

        part_sequences = {}
        for seq_name in all_part_names:
            if seq_name in self.available_sequences:
                part_sequences[seq_name] = self.available_sequences[seq_name].seq

        # Search plasmid files to find which contain each part
        part_to_plasmid = {}

        logger.info(f"Searching {len(plasmid_files)} plasmid files for part sequences...")

        for plasmid_file in plasmid_files:
            try:
                # Use parse to handle multiple records in a file
                for plasmid_record in SeqIO.parse(plasmid_file, "genbank"):
                    plasmid_seq = plasmid_record.seq.upper()

                    # Check each part sequence against this plasmid
                    for part_name, part_seq in part_sequences.items():
                        if part_name not in part_to_plasmid:  # Only map if not already found
                            part_seq_upper = part_seq.upper()
                            # Check both forward and reverse complement
                            if (part_seq_upper in plasmid_seq or
                                part_seq_upper.reverse_complement() in plasmid_seq):
                                part_to_plasmid[part_name] = plasmid_file.stem  # Just filename without extension
                                break  # Found the part, move to next part

            except Exception as e:
                logger.warning(f"Could not read {plasmid_file}: {str(e)}")
                continue
        
        # check for missing mappings
        missing_parts = all_part_names - set(part_to_plasmid.keys())

        if missing_parts:
            logger.error(f"Could not find plasmid containing: {', '.join(missing_parts)}")
            raise RuntimeError(f"Parts not found in any plasmids: {missing_parts}")
        

        # Get unique plasmid names for assembly
        required_plasmids = list(set(part_to_plasmid.values()))
        
        # Store results
        self.part_to_plasmid_mapping = part_to_plasmid
        self.plasmid_names = required_plasmids

        # Show mapping summary - too verbose, but useful for debugging
        # self.console.print(f"[green]Successfully mapped {len(all_part_names)} parts to {len(required_plasmids)} plasmids[/green]")

        # # Show detailed mapping
        # table = Table(title="Part to Plasmid Mapping")
        # table.add_column("Part Name", style="cyan")
        # table.add_column("Plasmid Name", style="yellow")

        # for part_name in sorted(all_part_names):
        #     table.add_row(part_name, part_to_plasmid[part_name])

        # self.console.print(table)

        return self.plasmid_names


    def gg_assembly(self, assembly_name: str) -> dc.Type2sRestrictionAssembly:
        """Pass plasmids to dnacauldron for Golden Gate assembly simulation.

        Loads plasmids from both public and private directories to create the repository.
        Sets self.assembly_sim = AssemblySimulation object.

        Args:
            assembly_name: name for plasmids in output

        Raises:
            ValueError: If no plasmids have been set
            ValueError: If no assemblies have been selected
            RuntimeError: If assembly simulation fails or produces unexpected results
        """
        from pathlib import Path
        

        # Construct both public and private plasmid paths
        public_plasmids = self.gg_plasmids

        try:
            relative_path = public_plasmids.relative_to("data")
            private_plasmids = Path("data/private") / relative_path
        except ValueError:
            private_plasmids = Path("data/private") / public_plasmids.name
        
        # Create repository and load records from both locations
        repository = dc.SequenceRepository()

        # Suppress BioPython FASTA comment deprecation warning
        # This warning is triggered by dnacauldron's use of BioPython's FASTA parser
        # with deprecated defaults, even when FASTA files don't contain comments
        with warnings.catch_warnings():
            warnings.filterwarnings(
                'ignore',
                category=BiopythonDeprecationWarning
            )

            # Load from public directory if it exists
            if public_plasmids.exists():
                repository.import_records(
                    folder=str(public_plasmids),
                    use_file_names_as_ids=False
                )

            # Load from private directory if it exists
            if private_plasmids.exists():
                repository.import_records(
                    folder=str(private_plasmids),
                    use_file_names_as_ids=False
                )

        self.repository = repository
        
        
        # Create Type 2 restriction assembly with the required plasmids
        assembly = dc.Type2sRestrictionAssembly(
            parts=self.plasmid_names,
            name=assembly_name,
            expected_constructs=len(self.assemblies_names),
            max_constructs=len(self.assemblies_names) + 1
        )
        self.assembly = assembly
        
        simulation = assembly.simulate(sequence_repository=repository)
        self.assembly_sim = simulation

        self._validate_assembly_results(simulation, assembly)
        
        return simulation


    def _validate_assembly_results(self, simulation, assembly) -> None:
        """Validate the assembly results from dnacauldron match the expected output"""

        # Check for errors
        if simulation.errors:
            for error in simulation.errors:
                logger.error(f"Assembly error: {error}")

        # Check for warnings
        if simulation.warnings:
            for warning in simulation.warnings:
                logger.warning(f"Assembly warning: {warning}")

        # Check number of constructs produced
        expected_constructs = len(self.assemblies_names)
        actual_constructs = len(simulation.construct_records)
        if actual_constructs != expected_constructs:
            logger.warning(f"Expected {expected_constructs} constructs, got {actual_constructs}")
            mixes = list(record["parts"] for record in simulation.compute_all_construct_data_dicts())
            if len(mixes) > 0:
                logger.warning(f"Mixtures found (enzyme: {assembly.enzyme}): {mixes}")
        else:
            logger.info(f"Generated {expected_constructs} constructs as expected using {assembly.enzyme}")

    def gg_save_output(self, output_path: str) -> None:
        """Saves the output of the assembly function, creating a new plasmid output file
        single assemblies, or a folder full of .gb files (no images) for multiplex assemblies.
        Takes output from simulation object to write a full report. also uses the user input
        from output_prefix to overwirte assembly names with input1, input2 etc.
        Outputs a graphical representations of multiplex assemblies.

        Args:
            Output_prefix: path to output folder,



        """
        reporter = dc.AssemblyReportWriter(include_part_records=False
                                           )
        self.assembly_sim.write_report(output_path, report_writer=reporter)



    def gg_instructions(self, output_path: str, assembly_name: str, liquid_handler: str = "human"):
        """Generate human-readable or machine-specific assembly instructions.

        Args:
            output_path: Path to output folder.
            assembly_name: Name prefix for output files.
            liquid_handler: Liquid handler to generate instructions for.
                One of 'janus', 'hamilton', 'epmotion', or 'human' (default: 'human').
        """
        all_construct_data = self.assembly_sim.compute_all_construct_data_dicts()
        for i, dict in enumerate(all_construct_data):
            # generate destination well, starts at A1 and fill horozontally (A1, A2, A3... etc ), for a library everything goes into one well
            if self.is_library:
                row_letter = 'A'
                col_number = '1'
            else:
                row_letter = chr(65 + ((i) // 12))  # A, B, C, etc.
                col_number = ((i) % 12) + 1        # 1, 2, 3, etc.
            dict["destination_well"] = f"{row_letter}{col_number}"
            parts = dict["parts"]
            wells = []
            for part in parts:
                loc = self._get_template_position(part)
                wells.append(loc)
            dict["wells"] = wells

        # Generate a dataframe with all the relavent information
        all_info_dataframe = pd.DataFrame(all_construct_data)

        instructions_dataframe = all_info_dataframe.explode("wells").reset_index(drop=True)

        if self.is_library:
            instructions_dataframe = instructions_dataframe.drop_duplicates(subset=["wells"]).reset_index(drop=True)

        liquid_handler = liquid_handler.strip().lower()

        if liquid_handler == "janus" or liquid_handler == "hamilton":
            # instructions_dataframe = all_info_dataframe.explode("wells").reset_index(drop=True)
            instructions_dataframe[['asperate_plate', 'asperate_well']] = pd.DataFrame(instructions_dataframe['wells'].to_list())
            instructions_dataframe = instructions_dataframe.drop('wells', axis = 1)
            janus_instructions = instructions_dataframe[['construct_id', 'asperate_plate', 'asperate_well', 'destination_well']].copy()
            janus_instructions['transfer_volume'] = 1
            janus_instructions['destination_plate'] = 'assembly plate' # need in increment this up by one ever 96 assemblies (will this ever happen? It's so much material!)
            janus_instructions = janus_instructions[['construct_id', 'asperate_plate', 'asperate_well', 'destination_plate', 'destination_well', 'transfer_volume']]
            janus_instructions.to_csv(f"{output_path}/{assembly_name}_worklist.csv", index = False)
            logger.info(f"Saved worklist to {output_path}/{assembly_name}_worklist.csv")

            # For single head machines you can use these columns to save tips. Different, v. complex implementation required for multi see CRVP paper by Wu et al 2025
            # instructions_dataframe = instructions_dataframe.sort_values(by = ['asperate_plate','asperate_well'])
            # janus_instructions['new_tip'] = (
            #     janus_instructions['asperate_well'] != janus_instructions['asperate_well'].shift()).map({True:'T', False: 'F'})
            # janus_instructions['drop_tip'] = (
            #     janus_instructions['asperate_well'] != janus_instructions['asperate_well'].shift(-1)).map({True:'T', False: 'F'})
            # janus_instructions.loc[0, 'new_tip'] = 'T'
            # janus_instructions.loc[len(janus_instructions)-1, 'drop_tip'] = 'T'

        elif liquid_handler == 'epmotion':
            header = [
                ['Labware', 'Src.Barcode', 'Src.List Name', 'Dest.Barcode', 'Dest.List name', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['Barcode ID', 'Labware', 'Source', 'Labware', 'Destination', 'Volume', 'Tool', 'Name']
            ]
            # instructions_dataframe = all_info_dataframe.explode("wells").reset_index(drop=True)
            instructions_dataframe[['Barcode ID', 'Source']] = pd.DataFrame(instructions_dataframe['wells'].to_list())
            instructions_dataframe = instructions_dataframe.drop('wells', axis = 1)
            # Map unique barcodes to sequential ID numbers
            unique_barcodes = instructions_dataframe['Barcode ID'].unique()
            barcode_to_labware = {barcode: i+1 for i, barcode in enumerate(unique_barcodes)}
            instructions_dataframe['Source_Labware'] = instructions_dataframe['Barcode ID'].map(barcode_to_labware)
            instructions_dataframe['Destination_labware'] = 1 # Supports a max of one assembly plate for now.
            instructions_dataframe['Tool'] = 'TS_50'# single channel only for cherry picking
            instructions_dataframe['Volume'] = 1
            instructions_dataframe['Name'] = ''
            epmotion_instructions = instructions_dataframe[['Barcode ID',
                                                            'Source_Labware',
                                                            'Source',
                                                            'Destination_labware',
                                                            'destination_well',
                                                            'Volume',
                                                            'Tool',
                                                            'Name'
                                                            ]]
            # Save output Note I need the csv wirter here - concatenating the header and df produces a misalignment.
            with open(f"{output_path}/{assembly_name}_epmotion_instructions.csv", 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(header)
                writer.writerows(epmotion_instructions.values)
            logger.info(f"Saved epMotion instructions to {output_path}/{assembly_name}_epmotion_instructions.csv")

        elif liquid_handler == 'human':
            volume_per_part = 1  # µL
            filename = f'{output_path}/{assembly_name}_human_instructions.txt'

            with open(filename, 'w') as f:
                f.write('Human Assembly Instructions\n')

                # Group by construct and destination well
                grouped = instructions_dataframe.groupby(['construct_id', 'destination_well'])

                for (construct, dest_well), group in grouped:
                    # Get construct info from first row (all same for this construct)
                    first_row = group.iloc[0]

                    # Header
                    f.write('='*70 + '\n')
                    f.write(f'CONSTRUCT: {construct:12s} ---> Destination: {dest_well}\n')
                    f.write('='*70 + '\n')
                    f.write(f'Total parts: {first_row["number_of_parts"]} | ')
                    f.write(f'Construct size: {first_row["construct_size"]} bp | ')
                    f.write(f'Volume: {volume_per_part} µL each\n')
                    f.write(f'Enzyme: {first_row["enzymes"][0]}\n\n')

                    # Parts checklist
                    f.write('PARTS TO ADD:\n')
                    parts_list = first_row['parts']  # Get the full parts list
                    for idx, (_, row) in enumerate(group.iterrows()):
                        plate, well = row['wells']
                        if plate is None:
                            plate = "not found"
                        if well is None:
                            well = "??"
                        part_name = parts_list[idx] if idx < len(parts_list) else 'part'

                        f.write(f'[ ] {part_name:15s} | {plate:15s} {well:3s} | {volume_per_part} µL\n')

                    # Master mix
                    f.write('\nMASTER MIX:\n')
                    f.write('[ ] 2.0 µL   10X T4 Ligase Buffer\n')
                    f.write(f'[ ] 2.0 µL   {first_row["enzymes"][0]} NEB master mix\n')
                    f.write(f'[ ] {(16 - (volume_per_part*len(group))):.1f} µL Water (20 µL total)\n')

                    f.write('\n[ ] Assembly complete\n')
                    f.write('\n\n')
                logger.info(f"Saved instructions to {output_path}/{assembly_name}_human_instructions.txt")




    def _get_template_position(self, template_name: str) -> tuple[Optional[str], Optional[str]]:
        """Find the plate and well position for a given template.

        First checks if the template is a contig/chromosome in the genome mapping file,
        then checks the template plates Excel file for individual templates.
        Searches both public and private template directories.

        Args:
            template_name: Name of the template or genome contig to locate

        Returns:
            Tuple containing:
                - Plate name/barcode (or None if not found)
                - Well position (or None if not found)
                Format of well position is standard 96-well notation (e.g., 'A1', 'H12')

        Raises:
            FileNotFoundError: If required mapping files aren't found
        """

        # Construct private template directory path
        private_template_folder = get_templates_path(private=True)

        # Check both public and private genome mapping files
        for search_dir in [self.template_folder, private_template_folder]:
            if not search_dir.exists():
                continue

            genome_map_file = search_dir / 'genome_well_mapping.tsv'
            if genome_map_file.exists():
                try:
                    genome_df = pd.read_csv(genome_map_file, sep='\t')

                    # Search through each row's contig list
                    for _, row in genome_df.iterrows():
                        contigs = [c.strip() for c in row['Contig_Names'].split(',')]
                        if template_name in contigs:
                            return row['Plate'], row['Well_Position']

                except Exception as e:
                    logger.warning(f"Error reading genome mapping file: {str(e)}")

        # If not found in genome mapping, check template plates Excel files
        for search_dir in [self.template_folder, private_template_folder]:
            if not search_dir.exists():
                continue

            template_excel = search_dir / 'TemPlates.xlsx'
            if not template_excel.exists():
                continue

            try:
                wb = openpyxl.load_workbook(template_excel)

                # Search each sheet (plate)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in range(3, 11):  # Rows 3-10 map to A-H
                        for col in range(2, 14):  # Columns 2-13 map to 1-12
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value == template_name:
                                # Convert Excel row/col to plate coordinates
                                well = f"{chr(row-3+65)}{col-1}"  # Convert 3->A, 4->B, etc.
                                return sheet, well

            except Exception as e:
                logger.warning(f"Error reading template plate map: {str(e)}")

        logger.warning(f"Template {template_name} not found in any mapping files")
        return None, None


    def find_templates(self, template_folder: Path) -> None:
        """Find template matches for each assembly component.

        Uses template functions from sequence_utils to identify potential
        templates for each sequence and rationalize the selections.

        Args:
            template_folder: Path to folder containing template files

        Raises:
            ValueError: If no sequences have been selected for assembly
        """
        if not self.assembly_sequences:
            raise ValueError("No sequences selected. Please select sequences first.")

        if not self.multiplex:
            self.template_dict = get_templates(self.assembly_sequences, str(template_folder))



