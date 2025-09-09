# Golden gate of level 1 plasmids using the MoClo YTK design 



# See LICENSE for full GpLv2 license. 

# This program is free software: you can redistribute it and/or modify 
# it under the terms of the GNU General Public License or 
# (at your option) any later version. 

# This program is distributed in the hope that it will be useful;, 
# but WITHOUT ANY WARRENTY; without even the implied warranty of 
# MERCHANTABILITY of FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details. 

# You should have received a copy of the GNU General Public License along
# with this program; if not, write to the Free Software Foundation, Inc., 
# 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA. 

# ===========================================================================

# src/pyeast/core/gg_lvl1.py
"""
Level 1 golden gate assemblies using the yeast Moclo standard from Lee et al 2015
"""

# ===========================================================================



from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from Bio import SeqIO
from Bio.SeqRecord import SeqRecord 
from Bio.Seq import Seq
from rich.console import Console
from rich.table import Table
from rich.progress import Progress
import click
from prompt_toolkit import PromptSession 
from prompt_toolkit.completion import WordCompleter
from prompt_toolkit.shortcuts import confirm
import io
from PIL import Image
import dnacauldron as dc
import pandas as pd
import openpyxl

from ..utils.sequence_utils import (
    load_sequences,
    get_templates,  
) 

from ..utils.visualisation import visualise_genbank, save_figure


class gg_lvl1Designer: 
    """A class for designing golden gate level 1 assemblies. 
    
    This class handles the design assemblies according to the 
    parameters in the yeast MoClo took kit laid out in Lee et al
    2015
    """
    #Flanking sequences specific to different part types, should specifiy part type as a string
    #'3a', '2', etc. and sequences as tuple of strings e.g. '1': ('5'seq', '3'seq')
    FLANKING_SEQs = {
        "1": ('CCCT','TTGC'),
        "2": ('AACG','ATAC'),
        "3": ('TATG', 'TAGG'),
        "3a": ('TATG', 'AAGA'),
        "3a'": ('TATG','TGCT'),
        "3b": ('TTCT', 'TAGG'),
        "3b'":('ACGA', 'TAGG'), 
        "4": ('ATCC', 'CGAC'),
        "4a": ('ATCC', 'ACCG'),
        "4b": ('TGGC', 'CGAC'), 
        "5": ('GCTG', 'ATGT'),
        "6": ('TACA', 'CTCA'),
        "7": ('GAGT', 'GGCT'), 
        "8": ('CCGA', 'GGGA'),
        "8a": ('CCGA', 'GTTA'), 
        "8b": ('CAAT', 'GGGA'), 
        }
    

    def __init__(self, 
                 gg_plasmids: Path = Path('data/gg plasmids/Yeast MoClo lvl 0'),
                 template_folder: Path = Path("data/templates")
                 ): 
        """Inintialise a new gg_lvl1 designer. 
        
        Args:
            gg_plasmids: Path to directory containing level 0 plasmid files (default: 'data/gg plasmids/Yeast MoClo lvl 0')
            """
        # File paths 
        self.gg_plasmids = gg_plasmids
        self.template_folder = template_folder
        
        # state storage
        self.console = Console()
        self.session = PromptSession() 
        self.multiplex = False
        self.available_sequences = {}     # All loaded sequences
        self.assembly_sequences = []      # Sequences in assembly order may be multiple assemblies for multiplex work (list of lists)
        self.assemblies_names = None      # Part names stored as strings
        self.template_dict = {}           # Template information
        self.repository = None            # dnacauldron object constructed from all lvl 0 parts plasmids for gg assembly
        self.assembly = None              # Assembly object from dnacauldron
        self.assembly_sim = None                # Simulation object from dnacauldron
        self.final_assembly = None        # Final assembled sequence(s)


    def load_and_get_sequences(self, directory: Path) -> None: 
        """"Loads seqeunces from a directory and store them for assembly. 
        
        Args: 
            directory: Path to a directory containing fasta files containing DNA components 
            as fasta files
            """
        self.available_sequences = load_sequences(directory)
        return self.available_sequences
    
    def print_sequence_grid(self, sequences: dict, title: str = "Available Sequences"):
        """Display available sequences in a formatted table"""
        table = Table(title=title)
        table.add_column("Name", style="cyan")
        table.add_column("Length", justify="right", style="green")
        table.add_column("Description", style="white")
        
        for name, seq in sequences.items():
            table.add_row(
                name,
                f"{len(seq)} bp",
                seq.description[len(name):150+len(name)] + "..." if len(seq.description)-len(name) > 149 else seq.description[len(name):]
            )
        
        self.console.print(table) 

    def get_assembly_order(self, sequences: Dict[str, SeqRecord]) -> List[list[str]]:
        """Get assembly order from user with autocomplete and mutliplex support 

        prompts the user to enter sequences for assembly with autocomplete. 
        Supports multiplex assemblies using '/' syntax to select multiple components 
        in a given position, or /allx to select all parts of a give type x. 
        
        Args: 
            sequences: Dictionary mapping sequence names to SeqRecord objects 

        Returns: 
            List[list[str]]: list of assemblies where each assembly is a list of sequence
                            names. For single assemblies returns [[seq1,seq2 ...]]
                            For multiplex assemblies returns [[seq1, seq2a,..], [seq1, seq2b...]]
        Raises: 
            click.Abort: if user cancels the operation
            ValueError If no valid sequences are available

        Examples:
        Single assembly: "promoter_1 gene_1 terminator_1" -> [["promoter_1", "gene_1", "terminator_1"]]
        Multiplex: "promoter_1 gene_1/gene_2 terminator_1" -> [["promoter_1", "gene_1", "terminator_1"], 
                                                               ["promoter_1", "gene_2", "terminator_1"]]
        All parts: "promoter_1 /all3 terminator_1" -> Multiple assemblies with type 3 parts
        """
        if not sequences: 
            raise ValueError("No sequences available for assembly")
        
        sequence_completer = WordCompleter(sequences.keys(), ignore_case=True)
        
        while True:
            try:
                self.console.print("\n[blue]Enter sequences to assemble (space-separated)[/blue]")
                self.console.print("[dim]Use TAB for autocompletion[/dim]")
                self.console.print("[dim]Remember that golden gate parts need to be assembled in order[/dim]")
                self.console.print("[blue]Separate components with / to multiplex, use /all[red]X[/red] to select all parts of type [red]x[/red]")
                
                user_input = self.session.prompt(
                    "Sequences: ",
                    completer=sequence_completer
                )
                
                selected = user_input.split()
                if not selected: 
                    self.console.print("[yellow]No sequences selected[/yellow]")
                    continue

                #validate selections before processing
                invalid_selections = []
                for selection in selected: 
                    if selection.startswith("/all"): 
                        part_type = selection[4:].strip()
                        matching_parts = [name for name in sequences.keys() if name.split("_")[0] == part_type]
                        if not matching_parts: 
                            invalid_selections.append(f"{selection} (no parts found with type '{part_type}')")

                    elif "/" in selection: 
                        part_names = selection.split('/')
                        for part_name in part_names: 
                            if part_name not in sequences: 
                                invalid_selections.append(f"{part_name} (from multplex selection '{selection})")
                    
                    else: 
                        if selection not in sequences: 
                            invalid_selections.append(selection)
                
                #send user back to select if there are invalid entries
                if invalid_selections: 
                    self.console.print(f"[red]Invalid selection(s): {','.join(invalid_selections)}[/red]")
                    continue
                
                # Process valid selections
                assemblies = [[]]
                for selection in selected: 
                    
                    #if the user has input /allx create new assemblies for each part of type x
                    if selection.startswith("/all"): 
                        self.multiplex = True # flag that this is now a multiplex assembly
                        part_type = selection[4:].strip() #part name should follow /all directly

                        multiplex_part_names = []
                        for part in sequences.keys(): 
                            if part.split("_")[0] == part_type: 
                                multiplex_part_names.append(part)
                        
                        print(f"{len(multiplex_part_names)} type {part_type} parts for multiplexing")
                        
                        #validate that the part name exists in the library 
                        if not multiplex_part_names: 
                            self.console.print(f"[red]No parts found with type {part_type}. Available types: {set(p.split('_')[0] for p in sequences.keys())}[/red]")
                      
                        #create new assemblies each with one of the multiplex parts appended
                        new_assemblies = []
                        for assembly in assemblies: 
                            for part in multiplex_part_names: 
                                new_assemblies.append(assembly + [part])
                        assemblies = new_assemblies 



                    elif "/" in selection: 
                        self.multiplex = True
                        part_type = selection.split("_")[0]
                          
                        part_names = selection.split("/")
                        print(f"{len(part_names)} type {part_type} parts for multiplexing")   

                        new_assemblies = []
                        for assembly in assemblies: 
                            for part in part_names: 
                                new_assemblies.append(assembly + [part])
                        assemblies = new_assemblies
                        # print(f"{len(assemblies)} multiplex assemblies with type {part_type} parts")
                                

                    else: #only one option selected, append to all assemblies in place
                        for assembly in assemblies: 
                            assembly.append(selection)

        
                #provide some feed back on how many constructs have been desinged in the combinatorial selection
                if self.multiplex:
                    print(f"{len(assemblies)} constructs for assembly")
                else: 
                    print("one construct for assembly")
                
                self.assemblies_names = assemblies
                return assemblies

            except Exception as e:
                print(f"Error: {str(e)}")
                raise click.Abort()
            
            except KeyboardInterrupt:
                if confirm("\nDo you want to exit?"):
                    raise click.Abort()
                continue  


    def get_plasmid_names(self) -> List[str]:  
        """this will map the sequences in self.assembly_sequences onto plasmids containing those sequences
        in the speficied directory, sets self.plasmid_names = list[str] doe dnacauldron 
        assembly
        
        Args: 
            plasmid_folder: Path to a directory containing level 0 plasmid files
        Returns: 
            List[str] of plasmid filenames (not paths) for dnacauldron assembly 
        Raises: 
            ValueError: If no assemblies have been selected
            FileNotFoundError: If plasmid folder doesn't exist
            RuntimeError: If parts cannot be mapped to plasmids
        """
        if not hasattr(self, 'assemblies_names') or not self.assemblies_names: 
            raise ValueError("No assemblies slected. Please run get_assembly_order first")
        
        if not hasattr(self, 'assemblies_names') or not self.assemblies_names: 
            raise ValueError("No assembly sequence available. Plaease run get_seq_records first.")
        
        if not self.gg_plasmids.exists(): 
            raise FileNotFoundError(F"Plasmid folder not found {self.gg_plasmids}")
        
        # Collect all the unique sequence names from all assemblies 
        all_part_names = set() 
        for assembly in self.assemblies_names: 
            all_part_names.update(assembly)

        part_sequences = {} 
        for seq_name in all_part_names: 
            if seq_name in self.available_sequences: 
                part_sequences[seq_name] = self.available_sequences[seq_name].seq
            
        # Search plasmid files to find which contain each part 
        part_to_plasmid = {}
        plasmid_files = list(self.gg_plasmids.glob('*.gb')) + list(self.gg_plasmids.glob('*.gbk'))


        self.console.print(f"[blue]Searching {len(plasmid_files)} plasmid files for part sequences...[/blue]")

        for plasmid_file in plasmid_files:
            try:
                # Use parse to handle multiple records in a file
                for plasmid_record in SeqIO.parse(plasmid_file, "genbank"):
                    plasmid_seq = plasmid_record.seq.upper()
                    
                    # Check each part sequence against this plasmid
                    for part_name, part_seq in part_sequences.items():
                        if part_name not in part_to_plasmid:  # Only map if not already found
                            part_seq_upper = part_seq.upper()
                            # Check both forward and reverse complement
                            if (part_seq_upper in plasmid_seq or 
                                part_seq_upper.reverse_complement() in plasmid_seq):
                                part_to_plasmid[part_name] = plasmid_file.stem  # Just filename without extension
                                break  # Found the part, move to next part
            
            except Exception as e: 
                self.console.print(f"[yellow]Warning could not read {plasmid_file}: {str(e)}[/yellow]")
                continue 

        # check for missing mappings
        missing_parts = all_part_names - set(part_to_plasmid.keys())

        if missing_parts: 
            self.console.print(f"[red]Could not find plasmid containsin {', '.join(missing_parts)}[/red]")
            raise RuntimeError(f"Parts not found in any plasmids: {missing_parts}")
        
        # Get unique plasmid names for assembly
        required_plasmids = list(set(part_to_plasmid.values()))

        # Store results 
        self.part_to_plasmid_mapping = part_to_plasmid 
        self.plasmid_names = required_plasmids

        # Show mapping summary
        self.console.print(f"[green]Successfully mapped {len(all_part_names)} parts to {len(required_plasmids)} plasmids[/green]")
        
        # # Show detailed mapping
        # table = Table(title="Part to Plasmid Mapping")
        # table.add_column("Part Name", style="cyan")
        # table.add_column("Plasmid Name", style="yellow")
        
        # for part_name in sorted(all_part_names):
        #     table.add_row(part_name, part_to_plasmid[part_name])
        
        # self.console.print(table)
        
        return self.plasmid_names
        

    def gg_assembly(self, assembly_name: str): 
        """Passes the set of plasmids in self.plasmid_names to dc.Type2RestrictionAssembly with 
        expected_constructs = len(self.assemblies_names), if there is warnings or errors it will 
        provide feedback on what went wrong and return the user to the sequence selection stage. 
        Sets self.final_final = AssemblySimulation object. Note this method doesn't know what the
        intention of the user is - that's restricted to the selection stage, but will check that 
        the outcome is roughly as expected. With this method the order of entry won't matter.
        
        Args: 
            assembly_name - name for plasmids in output

        Raises: 
            ValueError: If no plasmids have been set 
            ValueError: If no assemblies have been selected 
            RuntimeError: If assembly simulation fails or produces unexpected results
        """

        #Create repository from all available lvl 0 plasmids 
        repository = dc.SequenceRepository()
        repository.import_records(
            folder = str(self.gg_plasmids)
        )
        self.repository = repository 
        
        # Create Type 2 restriction assembly with the required plasmids 
        #  not sure if I'll need (all) the state storage for something or not
        assembly = dc.Type2sRestrictionAssembly(
            parts = self.plasmid_names,
            name = assembly_name,
            expected_constructs = len(self.assemblies_names), 
            max_constructs = len(self.assemblies_names) + 1
        )    
        self.assembly = assembly 

        simulation = assembly.simulate(sequence_repository = repository)
        self.assembly_sim = simulation

        self._validate_assembly_results(simulation, assembly)
        return simulation


    def _validate_assembly_results(self, simulation, assembly) -> None:
        """Validate the assembly results from dnacauldron match the expected output""" 
       
        # Check for errors 
        if simulation.errors:
            self.console.print("[red]Assembly errors detected:[/red]") 
            for error in simulation.errors: 
                self.console.print(f"[red] - {error}[/red]")
        
        # Check for warnings 
        if simulation.warnings: 
            self.console.print("[yellow]Assembly warnings:[/yellow]")
            for warning in simulation.warnings: 
                self.console.print(f"[yellow] - {warning}[/yellow]")

        # Check number of constructs produced
        expected_constructs = len(self.assemblies_names)
        actual_constructs = len(simulation.construct_records)
        if actual_constructs != expected_constructs: 
            self.console.print(f"[yellow]Expected {expected_constructs}, got {actual_constructs}[/yellow]")
        elif actual_constructs == expected_constructs: 
            self.console.print(f"generated {expected_constructs} constructs as expected using [blue]{assembly.enzyme}[/blue]")

    def gg_save_output(self, output_path: str) -> None: 
        """Saves the output of the assembly function, creating a new plasmid output file
        single assemblies, or a folder full of .gb files (no images) for multiplex assemblies. 
        Takes output from simulation object to write a full report. also uses the user input 
        from output_prefix to overwirte assembly names with input1, input2 etc. 
        Outputs a graphical representations of multiplex assemblies. 

        Args: 
            Output_prefix: path to output folder, 



        """
        reporter = dc.AssemblyReportWriter(include_part_records=False
                                           )
        self.assembly_sim.write_report(output_path, report_writer=reporter)
        
        

    def gg_instructions(self, output_path: str, assembly_name: str): 
        """Generates human readable and [optionally] machine specific instructions 
        this will just boil down to a table of what plasmids to include and where they are
        I will save this in the appropriate folder - but I'd like to make it possible to 
        run a gg_lvl1 with an option to start here from an existing folder full of plasmids 
        to be assembled - e.g. if you realize you want to use a different machine.
        this might actaully be clearer as a separate command simlar to batch.
        """
        all_construct_data = self.assembly_sim.compute_all_construct_data_dicts()
        for i, dict in enumerate(all_construct_data): 
            # generate destination well, starts at A1 and fill horozontally (A1, A2, A3... etc )
            row_letter = chr(65 + ((i) // 12))  # A, B, C, etc.
            col_number = ((i) % 12) + 1        # 1, 2, 3, etc.
            dict["destination_well"] = f"{row_letter}{col_number}"
            parts = dict["parts"]
            wells = []
            for part in parts: 
                loc = self._get_template_position(part)
                wells.append(loc)
            dict["wells"] = wells

        # Construct a big long list of instructions for a liquid handling robot for now (beta) this is specific to a janus robot. I'll we configure latter to support 
        # different instruments. 
        all_info_dataframe = pd.DataFrame(all_construct_data)
        instructions_dataframe = all_info_dataframe.explode("wells").reset_index(drop=True)
        instructions_dataframe[['asperate_plate', 'asperate_well']] = pd.DataFrame(instructions_dataframe['wells'].to_list())
        instructions_dataframe = instructions_dataframe.drop('wells', axis = 1)
        instructions_dataframe = instructions_dataframe.sort_values(by = ['asperate_plate','asperate_well'])
        liquid_handling_instructions = instructions_dataframe[['construct_id', 'asperate_plate', 'asperate_well', 'destination_well']].copy()
        liquid_handling_instructions['transfer_volume'] = 1
        liquid_handling_instructions['destination_plate'] = 'assembly plate'
        liquid_handling_instructions['new_tip'] = (
            liquid_handling_instructions['asperate_well'] != liquid_handling_instructions['asperate_well'].shift()).map({True:'T', False: 'F'})
        liquid_handling_instructions['drop_tip'] = (
            liquid_handling_instructions['asperate_well'] != liquid_handling_instructions['asperate_well'].shift(-1)).map({True:'T', False: 'F'})
        liquid_handling_instructions.loc[0, 'new_tip'] = 'T'
        liquid_handling_instructions.loc[len(liquid_handling_instructions)-1, 'drop_tip'] = 'T'
        

        #save output - I need to go back and change these to class variables when I add support for addition of other robots
        liquid_handling_instructions.to_csv(f"{output_path}/{assembly_name}_worklist.csv", index = False)
        self.console.print(f"Saved output to {output_path}/{assembly_name}_worklist.csv")
        #self.console.print(liquid_handling_instructions)
        
                                 

    def _get_template_position(self, template_name: str) -> Tuple[Optional[str], Optional[str]]:
        """Find the plate and well position for a given template.
        
        First checks if the template is a contig/chromosome in the genome mapping file,
        then checks the template plates Excel file for individual templates.
        
        Args:
            template_name: Name of the template or genome contig to locate
            
        Returns:
            Tuple containing:
                - Plate name/barcode (or None if not found)
                - Well position (or None if not found)
                Format of well position is standard 96-well notation (e.g., 'A1', 'H12')
                
        Raises:
            FileNotFoundError: If required mapping files aren't found
        """
        # First check genome mapping file for contigs
        genome_map_file = self.template_folder / 'genome_well_mapping.tsv'
        if genome_map_file.exists():
            try:
                genome_df = pd.read_csv(genome_map_file, sep='\t')
                
                # Search through each row's contig list
                for _, row in genome_df.iterrows():
                    contigs = [c.strip() for c in row['Contig_Names'].split(',')]
                    if template_name in contigs:
                        return row['Plate'], row['Well_Position']
                        
            except Exception as e:
                self.console.print(f"[yellow]Warning: Error reading genome mapping file: {str(e)}[/yellow]")
        
        # If not found in genome mapping, check template plates Excel file
        template_excel = self.template_folder / 'TemPlates.xlsx'
        if not template_excel.exists():
            raise FileNotFoundError(f"Template plate map not found: {template_excel}")
            
        try:
            wb = openpyxl.load_workbook(template_excel)
            
            # Search each sheet (plate)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in range(3, 11):  # Rows 3-10 map to A-H
                    for col in range(2, 14):  # Columns 2-13 map to 1-12
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value == template_name:
                            # Convert Excel row/col to plate coordinates
                            well = f"{chr(row-3+65)}{col-1}"  # Convert 3->A, 4->B, etc.
                            return sheet, well
                            
        except Exception as e:
            self.console.print(f"[yellow]Warning: Error reading template plate map: {str(e)}[/yellow]")
        
        # Template not found in either location
        self.console.print(f"[yellow]Warning: Template {template_name} not found in any mapping files[/yellow]")
        return None, None


    def find_templates(self, template_folder: Path) -> None:
        """Find template matches for each assembly component.
        
        Uses template functions from sequence_utils to identify potential
        templates for each sequence and rationalize the selections.
        
        Args:
            template_folder: Path to folder containing template files
            
        Raises:
            ValueError: If no sequences have been selected for assembly
        """
        if not self.assembly_sequences:
            raise ValueError("No sequences selected. Please select sequences first.")
        
        if not self.multiplex: 
            self.template_dict = get_templates(self.assembly_sequences, str(template_folder))



    