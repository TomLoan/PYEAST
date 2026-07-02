# Copyright ANU 2025. Thomas Loan
# See LICENSE for full GpLv2 license.

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License or
# (at your option) any later version.

# This program is distributed in the hope that it will be useful;,
# but WITHOUT ANY WARRENTY; without even the implied warranty of
# MERCHANTABILITY of FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License along
# with this program; if not, write to the Free Software Foundation, Inc.,
# 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.

# ===========================================================================

# src/pyeast/core/batch.py

"""
Batch assembly designer for high-throughput DNA assembly in S. cerevisiae.

This module provides tools for organizing multiple DNA assemblies into efficient
batches for parallel processing.
"""

# ===========================================================================





import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from Bio import SeqIO
from Bio.Seq import Seq
from Bio.SeqFeature import CompoundLocation
from Bio.SeqRecord import SeqRecord
from tabulate import tabulate

logger = logging.getLogger(__name__)

from pyeast.utils.path_utils import get_output_path, get_primers_path, get_templates_path
from pyeast.utils.primer_utils import get_primer_locations, rationalize_primers
from pyeast.utils.sequence_utils import get_templates, rationalize_templates


class BatchDesigner:
    """Designer class for batch DNA assembly experiments.

    This class handles the organization and planning of multiple DNA assemblies
    in parallel, including primer design, template selection, and generating
    machine-readable instructions for liquid handling robots.
    """
    EPMOTION_HEADER = [
                ['Labware', 'Src.Barcode', 'Src.List Name', 'Dest.Barcode', 'Dest.List name', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['', '', '', '', '', '', '', ''],
                ['Barcode ID', 'Labware', 'Source', 'Labware', 'Destination', 'Volume', 'Tool', 'Name']
            ]
    JANUS_HEADER = [['construct_id', 'asperate_plate', 'asperate_well', 'destination_plate', 'destination_well', 'transfer_volume']]

    def __init__(self,
                 reuse_limit = 5,
                 batch_size: int = 96,
                 primer_folder: Optional[Path] = None,
                 template_folder: Optional[Path] = None,
                 output_folder: Optional[Path] = None):
        """Initialize BatchDesigner with specified parameters.

        Args:
            reuse_limit: Maximum times a PCR product can be reused (default: 5)
            batch_size: Maximum number of PCRs per batch (default: 96)
            primer_folder: Path to folder containing primer files (default: data/primers)
            template_folder: Path to folder containing template files (default: data/templates)
            output_folder: Path to store generated GenBank files (default: output)
        """
        # Core parameters
        self.batch_size = batch_size

        # File paths - resolve defaults if not provided
        self.primer_folder = primer_folder if primer_folder is not None else get_primers_path()
        self.template_folder = template_folder if template_folder is not None else get_templates_path()
        self.output_folder = output_folder if output_folder is not None else get_output_path()

        # State storage
        self.available_constructs = {}  # GenBank files loaded
        self.selected_constructs = {}   # Constructs chosen for assembly
        self.validation_errors = []     # Issues with Genbank annotations and primers
        self.assembly_requirements = {}   # Componants and primers extracted from GenBank
        self.primer_locations = ()      # Primer locations ({primers_found}, {missing_primers})
        self.template_matches = {}      # Template locations
        self.assembly_instructions = []
        self.pcr_reactions = {}
        self.reuse_limit = reuse_limit
        self.batched_reactions = []
        self.output_folder.mkdir(parents=True, exist_ok=True)

    def load_constructs(self) -> None:
        """Load available constructs from output directory and subfolders.

        Searches for .gb files in:
        1. Root output directory (legacy files)
        2. Subfolders within output directory (new organized structure)

        Only loads constructs suitable for batch processing (from tar/integrate commands).
        Automatically excludes cassettes from replace/delete commands and gg assemblies.
        """
        VALID_DEFINITIONS = {
            "Plasmid assembled by TAR cloning simulation",
            "Assembled sequence for genomic integration"
        }
        self.available_constructs = {}

        # First, check for legacy files in the root output directory
        legacy_count = 0
        skipped_count = 0
        for filename in self.output_folder.glob("*.gb"):
            try:
                record = SeqIO.read(filename, "genbank")

                # Check if this is a valid target for batch processing
                description = record.description
                if description not in VALID_DEFINITIONS:
                    skipped_count += 1
                    # self.console.print(f"[dim]Skipped {filename.stem} (not a TAR/integrate output)[/dim]")
                    continue
                ## Use filename without extension as key (simple name only)
                self.available_constructs[filename.stem] = record
                record.annotations['source_folder'] = 'output/'
                record.annotations['source_type'] = 'TAR' if 'TAR' in description else 'integrate'
                legacy_count += 1
            except Exception as e:
                logger.error(f"Error reading {filename}: {str(e)}")

        # if legacy_count > 0:
        #     self.console.print(f"[yellow]Found {legacy_count} legacy construct(s) in root output directory[/yellow]")
        # if skipped_count > 0:
        #     self.console.print(f"[dim]Skipped {skipped_count} file(s) (not TAR/integrate outputs)[/dim]")

        # Then, check subfolders for new organized structure
        subfolder_count = 0
        for subfolder in self.output_folder.iterdir():
            if not subfolder.is_dir():
                continue

            # Count .gb files in this subfolder
            gb_files = list(subfolder.glob("*.gb"))
            if not gb_files:
                continue

            for filename in gb_files:
                try:
                    record = SeqIO.read(filename, "genbank")

                    # Check if this is a valid target for batch processing
                    description = record.description
                    # print(description)
                    if description not in VALID_DEFINITIONS:
                          skipped_count += 1
                          continue

                    # Use ONLY the filename as key (no folder path)
                    # This simplifies selection - user just types the construct name
                    construct_name = filename.stem

                    # Check for naming conflicts
                    if construct_name in self.available_constructs:
                        logger.warning(
                            f"Multiple constructs named '{construct_name}' found. "
                            f"Using the one from {subfolder.name}/"
                        )

                    self.available_constructs[construct_name] = record

                    # Store source folder for display and reference
                    record.annotations['source_folder'] = subfolder.name
                    record.annotations['source_type'] = 'TAR' if 'TAR' in description else 'integrate'
                    subfolder_count += 1

                except Exception as e:
                    logger.error(f"Error reading {filename}: {str(e)}")

        if subfolder_count > 0:
            logger.info(f"Found {subfolder_count} construct(s) in subfolders")

        if not self.available_constructs:
            raise ValueError(
                "No valid GenBank files found for batch processing.\n"
                "Ensure Output directory is correct. \n" 
                "Make sure you've run 'tar' or 'integrate' commands first to generate constructs."
            )

        logger.info(f"Loaded {len(self.available_constructs)} total construct(s)")

    def validate_constructs(self) -> None:
        """
        Validate all selected constructs for assembly compatibility.

        Checks that each construct:
        - Has a topology annotation
        - Has at least one part (PYEAST_component or misc_feature)
        - Each part has proper primer annotations:
            - Forward primer at start (strand=1)
            - Reverse primer at end (strand=-1)

        Raises:
            ValueError: If no constructs have been selected
            ValueError: If any validation errors are found
        """
        if not self.selected_constructs:
            raise ValueError("No constructs selected. Run get_selections first.")

        self.validation_errors = []

        for name, record in self.selected_constructs.items():
            # Check topology
            if 'topology' not in record.annotations:
                self.validation_errors.append(f"{name}: Missing topology annotation")
                continue

            is_circular = record.annotations['topology'].lower() == 'circular'

            # Get parts
            pyeast_parts = [f for f in record.features if f.type == "PYEAST_component"]
            parts = pyeast_parts if pyeast_parts else [f for f in record.features if f.type == "misc_feature"]
            if not parts:
                self.validation_errors.append(f"{name}: No parts (misc_feature) found")
                continue

            # Validate each part's primers
            for part in parts:
                part_label = part.qualifiers.get('label', ['Unlabeled'])[0]
                part_start = int(part.location.start)
                part_end = int(part.location.end)

                # Find primers that could overlap this part's boundaries
                primers = [p for p in record.features if p.type == "primer_bind"]

                # Check start primers
                start_primers = []
                for primer in primers:
                    # Handle compound locations for circular sequences
                    if isinstance(primer.location, CompoundLocation):
                        # For primers spanning origin, check if any part overlaps our target
                        for sublocation in primer.location.parts:
                            if (int(sublocation.start) <= part_start <= int(sublocation.end) or
                                (is_circular and int(sublocation.end) < int(sublocation.start))):
                                if primer.location.strand == 1:
                                    start_primers.append(primer)
                    else:
                        # Normal location check
                        if (int(primer.location.start) <= part_start <= int(primer.location.end) and
                            primer.location.strand == 1):
                            start_primers.append(primer)

                if not start_primers:
                    self.validation_errors.append(
                        f"{name}: Part '{part_label}' missing forward (strand=1) primer at start"
                    )

                # Check end primers
                end_primers = []
                for primer in primers:
                    # Handle compound locations for circular sequences
                    if isinstance(primer.location, CompoundLocation):
                        # For primers spanning origin, check if any part overlaps our target
                        for sublocation in primer.location.parts:
                            if (int(sublocation.start) <= part_end <= int(sublocation.end) or
                                (is_circular and int(sublocation.end) < int(sublocation.start))):
                                if primer.location.strand == -1:
                                    end_primers.append(primer)
                    else:
                        # Normal location check
                        if (int(primer.location.start) <= part_end <= int(primer.location.end) and
                            primer.location.strand == -1):
                            end_primers.append(primer)

                if not end_primers:
                    self.validation_errors.append(
                        f"{name}: Part '{part_label}' missing reverse (strand=-1) primer at end"
                    )

                # Validate part sequence
                if not part.extract(record.seq):
                    self.validation_errors.append(
                        f"{name}: Part '{part_label}' has invalid sequence"
                    )

        if self.validation_errors:
            raise ValueError("Validation failed with the following errors:\n" +
                            "\n".join(f"  • {e}" for e in self.validation_errors))


    def extract_components_and_primers(self, record: SeqRecord) -> dict:
        """Extract component and primer information from GenBank annotations."""
        components = []

        # First extract all components
        for feature in record.features:
            if feature.type == "PYEAST_component":
                if "label" not in feature.qualifiers:
                    raise ValueError(f"Component at position {feature.location} missing label")

                component_seq = feature.extract(record.seq)
                component = {
                    'name': feature.qualifiers['label'][0],
                    'sequence': str(component_seq),
                    'start': int(feature.location.start),
                    'end': int(feature.location.end),
                    'forward_primer': None,
                    'reverse_primer': None
                }
                components.append(component)

            elif feature.type == "misc_feature" and not ('PYEAST_component' in [f.type for f in record.features]):
                if "label" not in feature.qualifiers:
                    raise ValueError(f"Component at position {feature.location} missing label")

                component_seq = feature.extract(record.seq)
                component = {
                    'name': feature.qualifiers['label'][0],
                    'sequence': str(component_seq),
                    'start': int(feature.location.start),
                    'end': int(feature.location.end),
                    'forward_primer': None,
                    'reverse_primer': None
                }
                components.append(component)

        if not components:
            raise ValueError("No components (PYEAST_component or misc_features) found in record")

        # Sort components by position
        components.sort(key=lambda x: x['start'])

        # For each component, find its primers
        for component in components:
            start_primers = []  # Primers overlapping component start
            end_primers = []    # Primers overlapping component end

            for feature in record.features:
                if feature.type != "primer_bind" or "label" not in feature.qualifiers:
                    continue

                #primer_seq = feature.extract(record.seq)

                if isinstance(feature.location, CompoundLocation):
                    primer_start = int(list(feature.location)[0])
                    primer_end = int(list(feature.location)[-1])

                else:
                    primer_start = int(feature.location.start)
                    primer_end = int(feature.location.end)

                # Determine if forward or reverse from strand
                is_forward = feature.location.strand >= 0  # 1 or None is forward, -1 is reverse

                # For circular sequences, we need to check if the primer spans the origin
                if 'topology' in record.annotations and record.annotations['topology'].lower() == 'circular':
                    seq_length = len(record.seq)

                    # Adjust component positions for end-spanning cases
                    comp_start = component['start']
                    comp_end = component['end']

                    # Check for overlaps considering circular nature
                    overlaps_start = (
                        (primer_start <= comp_start <= primer_end) or
                        (primer_start >= primer_end and comp_start == 0)
                    )
                    overlaps_end = (
                        (primer_start <= comp_end <= primer_end and primer_end <= seq_length - 50) or
                        (comp_end == seq_length and primer_end >= (seq_length - 50) )
                    )
                else:
                    # Linear sequence - simple overlap check
                    overlaps_start = primer_start <= component['start'] <= primer_end
                    overlaps_end = primer_start <= component['end'] <= primer_end

                if overlaps_start or overlaps_end:
                    primer_seq = feature.extract(record.seq)
                    primer_info = {
                        'name': feature.qualifiers['label'][0],
                        'sequence': str(primer_seq),
                        'position': (primer_start, primer_end),
                        'is_forward': is_forward
                    }

                    if overlaps_start:
                        start_primers.append(primer_info)
                    if overlaps_end:
                        end_primers.append(primer_info)

            # Assign primers to component
            for primer in start_primers:
                if primer['is_forward']:
                    component['forward_primer'] = primer

            for primer in end_primers:
                if not primer['is_forward']:
                    component['reverse_primer'] = primer
            component["PCR_length"] = len(component["sequence"]) + 50

            if not component['forward_primer'] or not component['reverse_primer']:
                raise ValueError(f"Component {component['name']} missing primers")

        return components

    def process_selected_constructs(self) -> None:
        """Process all selected constructs to extract their assembly requirements.

        This function:
        1. Iterates through selected constructs
        2. Extracts components, primers, and sequences from each
        3. Stores all information needed for assembly instructions

        The assembly_requirements dict will be structured as:
        {
            'construct_name': {
                'record': SeqRecord,              # Original record for reference
                'topology': 'circular' or 'linear',
                'components': [
                    {
                        'name': str,              # Component name
                        'sequence': str,          # Component sequence
                        'start': int,             # Start position
                        'end': int,               # End position
                        'forward_primer': {
                            'name': str,          # Primer name
                            'sequence': str,      # Primer sequence
                            'position': (int, int) # Primer location
                        },
                        'reverse_primer': {
                            'name': str,
                            'sequence': str,
                            'position': (int, int)
                        }
                    },
                    ...
                ]
            }
        }

        Raises:
            ValueError: If no constructs have been selected
            ValueError: If assembly requirement extraction fails for any construct
        """
        if not self.selected_constructs:
            raise ValueError("No constructs selected. Run get_selections first.")

        self.assembly_requirements = {}
        processing_errors = []

        for name, record in self.selected_constructs.items():
            try:
                # Extract components and primers
                components = self.extract_components_and_primers(record)

                # Get topology from record annotations
                topology = record.annotations.get('topology', '').lower()
                if topology not in ['circular', 'linear']:
                    raise ValueError(f"Invalid topology annotation in {name}")

                # Store everything in class state
                self.assembly_requirements[name] = {
                    'record': record,             # Keep original record for reference
                    'topology': topology,
                    'components': components      # Contains all component and primer info
                }

                logger.info(f"Processed {name}: {len(components)} components, {topology} topology")

            except Exception as e:
                # Collect errors but continue processing other constructs
                error_msg = f"Error processing {name}: {str(e)}"
                processing_errors.append(error_msg)
                logger.error(error_msg)

        # If any constructs failed processing, raise error with details
        if processing_errors:
            raise ValueError(
                "Failed to process some constructs:\n" +
                "\n".join(f"  • {e}" for e in processing_errors)
            )

    def find_primers_and_templates(self) -> None:
        """Find primer locations, template matches, and track required PCR reactions.

        This function:
        1. Identifies all PCR reactions needed across assemblies
        2. Tracks reaction usage counts for batch planning
        3. Finds primer locations and template matches
        4. Rationalizes selections considering reuse limits

        Raises:
            ValueError: If assembly requirements haven't been processed
        """
        if not hasattr(self, 'assembly_requirements'):
            raise ValueError("No assembly requirements found. Run process_selected_constructs first.")

        # Track all required PCR reactions
        self.pcr_reactions = {}  # Dictionary to store all unique PCR reactions

        def get_homology_regions(component_seq: str, f_primer_seq: str, r_primer_seq: str) -> tuple[str, str]:
            """Find homology regions by checking where primers match component sequence.

            Args:
                component_seq: The component sequence
                f_primer_seq: Forward primer sequence
                r_primer_seq: Reverse primer sequence

            Returns:
                Tuple containing:
                    - Forward primer homology sequence
                    - Reverse primer homology sequence (in forward orientation)
            """
            # Forward primer - check for match at start of component
            f_homology = None
            for i in range(len(f_primer_seq)):
                annealing = f_primer_seq[i:]
                if component_seq.startswith(annealing):
                    f_homology = f_primer_seq[:i]
                    break

            if f_homology is None:
                raise ValueError("Could not find forward primer match in component")

            # For reverse primer:
            # 1. Get reverse complement of component end
            # 2. Search from 3' end of primer (right side)
            comp_end_rc = str(Seq(component_seq).reverse_complement())
            r_homology = None

            for i in range(len(r_primer_seq)):
                annealing = r_primer_seq[-i:] if i > 0 else r_primer_seq
                if comp_end_rc.startswith(annealing):
                    r_homology = r_primer_seq[:-i] if i > 0 else ''
                    r_homology = Seq(r_homology)
                    r_homology = str(r_homology.reverse_complement())
                    break

            if r_homology is None:
                raise ValueError("Could not find reverse primer match in component")

            return f_homology, r_homology

        for construct_name, construct_info in self.assembly_requirements.items():
            construct_reactions = []  # Store reactions for this construct

            for component in construct_info['components']:
                # Get primer sequences
                f_primer_seq = str(component['forward_primer']['sequence'])
                r_primer_seq = str(component['reverse_primer']['sequence'])
                component_seq = str(component['sequence'])

                # Find homology regions
                f_homology, r_homology = get_homology_regions(component_seq, f_primer_seq, r_primer_seq)

                # Construct complete PCR product sequence
                complete_product = f_homology + component_seq + r_homology

                # Create a reaction identifier that includes the complete product
                reaction_id = f"{complete_product}"

                # Store reaction details if we haven't seen it before
                if reaction_id not in self.pcr_reactions:
                    self.pcr_reactions[reaction_id] = {
                        'component_name': component['name'],
                        'component_sequence': component_seq,
                        'PCR_length': len(complete_product),
                        'complete_product': complete_product,
                        'forward_primer': {
                            'name': component['forward_primer']['name'],
                            'sequence': f_primer_seq,
                            'homology': f_homology
                        },
                        'reverse_primer': {
                            'name': component['reverse_primer']['name'],
                            'sequence': r_primer_seq,
                            'homology': r_homology
                        },
                        'used_in_constructs': [],
                        'total_uses': 0,
                        'needs_repeats': False
                    }

                # Track usage
                self.pcr_reactions[reaction_id]['used_in_constructs'].append(construct_name)
                self.pcr_reactions[reaction_id]['total_uses'] += 1

                # Check if we've exceeded reuse limit
                if self.pcr_reactions[reaction_id]['total_uses'] > self.reuse_limit:
                    self.pcr_reactions[reaction_id]['needs_repeats'] = True
                    total_uses = self.pcr_reactions[reaction_id]['total_uses']
                    # calculate the number of additional wells beyond the first required
                    num_repeats_needed = (total_uses - 1)//self.reuse_limit
                    self.pcr_reactions[reaction_id]['num_repeats_needed'] = num_repeats_needed
                else:
                    self.pcr_reactions[reaction_id]['num_repeats_needed'] = 0

                construct_reactions.append(reaction_id)

            # Store reactions needed for this construct
            self.assembly_requirements[construct_name]['required_reactions'] = construct_reactions

        # Now collect all unique primers for location finding
        all_primers = {}
        for reaction in self.pcr_reactions.values():
            all_primers[reaction['forward_primer']['name']] = reaction['forward_primer']['sequence']
            all_primers[reaction['reverse_primer']['name']] = reaction['reverse_primer']['sequence']

        # Find primer locations
        self.primers_found, self.missing_primers = get_primer_locations(
            all_primers,
            str(self.primer_folder)
        )

        # Collect components for template finding
        all_components = []
        for reaction in self.pcr_reactions.values():
            component_record = SeqRecord(
                Seq(reaction['component_sequence']),
                id=reaction['component_name'],
                name=reaction['component_name']
            )
            all_components.append(component_record)

        # Find template matches
        self.template_matches = get_templates(all_components, str(self.template_folder))
        #self.console.print(self.primers_found)
        # Rationalize selections
        self.rationalized_primers = rationalize_primers(
            self.primers_found,
            self.missing_primers
        )

        self.rationalized_templates = rationalize_templates(self.template_matches)

        total_reactions = len(self.pcr_reactions)
        repeated_reactions = sum(r['num_repeats_needed'] for r in self.pcr_reactions.values())
        logger.info(f"PCR reactions: {total_reactions} unique, {repeated_reactions} needing repeats")
        logger.info(f"Primers: {len(all_primers)} total, {len(self.primers_found)} found, {len(self.missing_primers)} to order")
        template_count = sum(1 for templates in self.template_matches.values()
                            if templates[0] != "Not found")
        no_template_count = sum(1 for template in self.template_matches.values() if template == "Not found")
        logger.info(f"Templates: {template_count} found, {no_template_count} not found")

    def organize_pcr_batches(self) -> None:
        """Organize PCR reactions into efficient batches.

        This function:
        1. Groups PCR reactions into batches of batch_size (default 96)
        2. Keeps reactions for the same construct together when possible
        3. Handles arbitarary numbers of repeated reactions needed due to reuse limits
        4. Maintains order from user's construct selection

        The batched_reactions list will contain dictionaries with structure:
        {
            'batch_number': int,
            'reactions': [
                {
                    'reaction_id': str,
                    'component_name': str,
                    'forward_primer': dict,
                    'reverse_primer': dict,
                    'template': str,
                    'constructs': List[str],
                    'num_repeats_needed': int
                },
                ...
            ],
            'constructs_completed': List[str]  # Constructs that can be assembled from this batch
        }

        Raises:
            ValueError: If PCR reactions haven't been identified
            ValueError: If required templates haven't been rationalized
        """
        if not hasattr(self, 'pcr_reactions'):
            raise ValueError("No PCR reactions found. Run find_primers_and_templates first.")
        if not self.rationalized_templates:
            raise ValueError("Template selections not rationalized.")

        self.batched_reactions = []
        current_batch = {
            'batch_number': 1,
            'reactions': [],
            'constructs_completed': []
        }

        # Track which reactions have been added
        processed_reactions = set()

        # Process constructs in original selection order
        for construct_name in self.assembly_requirements.keys():
            construct_info = self.assembly_requirements[construct_name]
            required_reactions = construct_info['required_reactions']

            # Check if all reactions for this construct will fit in current batch
            reactions_needed = []
            for reaction_id in required_reactions:
                reaction = self.pcr_reactions[reaction_id]
                # Count original reaction
                if reaction_id not in processed_reactions:
                    reactions_needed.append(reaction_id)
                # Count all repeats if needed
                num_repeats = reaction['num_repeats_needed']
                for repeat_num in range(1, num_repeats + 1):
                    repeat_id = f'{reaction_id}_repeat{repeat_num}'
                    if repeat_id not in processed_reactions:
                        reactions_needed.append(repeat_id)

            # If current batch would overflow, start a new one
            if len(current_batch['reactions']) + len(reactions_needed) > self.batch_size:
                self.batched_reactions.append(current_batch)
                current_batch = {
                    'batch_number': len(self.batched_reactions) + 1,
                    'reactions': [],
                    'constructs_completed': []
                }

            # Add reactions for this construct
            construct_complete = True
            for reaction_id in required_reactions:
                reaction = self.pcr_reactions[reaction_id]

                # Add original reaction if not already processed
                if reaction_id not in processed_reactions:
                    product_length = reaction['PCR_length']
                    reaction_info = {
                        'reaction_id': reaction_id,
                        'component_name': reaction['component_name'],
                        'forward_primer': reaction['forward_primer'],
                        'reverse_primer': reaction['reverse_primer'],
                        'template': self.rationalized_templates[reaction['component_name']],
                        'constructs': reaction['used_in_constructs'],
                        'is_repeat': False,
                        'repeat_number': 0,
                        'product_length': product_length
                    }
                    current_batch['reactions'].append(reaction_info)
                    processed_reactions.add(reaction_id)

                # Add all repeat reactions if needed
                num_repeats = reaction['num_repeats_needed']
                for repeat_num in range(1, num_repeats+1):
                    repeat_id = f"{reaction_id}_repeat{repeat_num}"
                    if repeat_id not in processed_reactions:
                        repeat_info = {
                            'reaction_id': repeat_id,
                            'component_name': reaction['component_name'],
                            'forward_primer': reaction['forward_primer'],
                            'reverse_primer': reaction['reverse_primer'],
                            'template': self.rationalized_templates[reaction['component_name']],
                            'constructs': reaction['used_in_constructs'],
                            'is_repeat': True,
                            'repeat_number': repeat_num,
                            'product_length' : product_length
                        }
                        current_batch['reactions'].append(repeat_info)
                        processed_reactions.add(repeat_id)

            # If we got all reactions for this construct, mark it as complete
            if construct_complete:
                current_batch['constructs_completed'].append(construct_name)

        # Add final batch if not empty
        if current_batch['reactions']:
            self.batched_reactions.append(current_batch)

        for batch in self.batched_reactions:
            logger.info(
                f"Batch {batch['batch_number']}: {len(batch['reactions'])} reactions, "
                f"completes: {', '.join(batch['constructs_completed'])}"
            )

    def generate_human_instructions(self, output_prefix: str) -> None:
        """Generate human-readable assembly instructions and save to files.

        Creates a table of instructions for each batch and saves to files:
        - {output_prefix}_batch_instructions.tsv: Main assembly instructions
        - {output_prefix}_missing_primers.tsv: Primers that need ordering

        Args:
            output_prefix: Prefix for output files (path + base filename)

        Raises:
            ValueError: If reactions haven't been batched
        """
        if not hasattr(self, 'batched_reactions'):
            raise ValueError("No batched reactions found. Run organize_pcr_batches first.")

        self.human_instructions = []

        # Add header
        self.human_instructions.append([
            "Batch", "Component", "Constructs",
            "F_Name", "F_Plate", "F_Well",
            "R_Name", "R_Plate", "R_Well",
            "Template", "Size", "Repeat?"
        ])

        # Generate instructions for each batch
        for batch in self.batched_reactions:
            batch_num = batch['batch_number']

            for reaction in batch['reactions']:
                # Get primer locations from rationalized primers
                f_primer_info = self.rationalized_primers[reaction['forward_primer']['name']]
                r_primer_info = self.rationalized_primers[reaction['reverse_primer']['name']]

                # Create instruction row
                row = [
                    batch_num,
                    reaction['component_name'],
                    ", ".join(reaction['constructs']),
                    reaction['forward_primer']['name'],
                    f_primer_info['Location'],
                    f_primer_info['Position'],
                    reaction['reverse_primer']['name'],
                    r_primer_info['Location'],
                    r_primer_info['Position'],
                    reaction['template'],
                    reaction['product_length'],
                    "Yes" if reaction['is_repeat'] else "No"
                ]

                self.human_instructions.append(row)

        # Save instructions to TSV
        instructions_file = f"{output_prefix}_batch_instructions.tsv"
        with open(instructions_file, 'w') as f:
            for row in self.human_instructions:
                f.write("\t".join(str(x) for x in row) + "\n")

        # Save missing primers to TSV
        if self.missing_primers:
            missing_primers_file = f"{output_prefix}_missing_primers.tsv"
            with open(missing_primers_file, 'w') as f:
                f.write("Name\tSequence\n")
                for name, info in self.missing_primers.items():
                    f.write(f"{name}\t{info[0]['sequence']}\n")
            logger.info(f"Missing primers: {missing_primers_file}")

        logger.info(f"Assembly instructions: {instructions_file}")

    def generate_assembly_groups(self, output_prefix: str) -> None:
        """Generate instructions for combining PCR products into final assemblies.

        Creates tables showing which PCR products to combine for each construct
        within each batch. Saves to file and displays in terminal.

        Args:
            output_prefix: Prefix for output files (path + base filename)

        Raises:
            ValueError: If batched reactions haven't been generated
        """
        if not hasattr(self, 'batched_reactions'):
            raise ValueError("No batched reactions found. Run organize_pcr_batches first.")

        # Build lookup for well positions from batched reactions
        # Key: (batch_num, complete_product) -> well_position
        # reaction_id_with_suffix can be "original_id", "original_id_repeat1", "original_id_repeat2", etc.
        reaction_wells = {}

        # Process each batch to assign wells to reactions
        for batch in self.batched_reactions:
            batch_num = batch['batch_number']
            for reaction in batch['reactions']:
                reaction_id = reaction['reaction_id']
                # Get complete product sequence from pcr_reactions
                # complete_product = self.pcr_reactions[reaction_id]['complete_product']
                # Store well with full reaction_id (including _repeatN suffix if present)
                well_key = (batch_num, reaction_id)
                # If this exact product already has a well, reuse it
                if well_key not in reaction_wells:
                    reaction_wells[well_key] = f"Well {len(reaction_wells) + 1}"

        # Track usage count per reaction across all constructs in order
        reaction_usage_count = {} # reaction_id -> current usage count

        assembly_groups = []
        # Add header
        assembly_groups.append([
            "Batch", "Construct", "Required_Wells", "Topology"
        ])

        # Process each construct that can be completed in this batch
        for batch in self.batched_reactions:
            batch_num = batch['batch_number']

            # Process each construct that can be completed in this batch
            for construct_name in batch['constructs_completed']:
                construct_info = self.assembly_requirements[construct_name]

                # Get the required reactions for this construct
                #required_reactions = []
                required_wells = []

                for reaction_id in construct_info['required_reactions']:
                    # Initialize usage count if not seen before
                    if reaction_id not in reaction_usage_count:
                        reaction_usage_count[reaction_id] = 0
                    reaction_usage_count[reaction_id] += 1
                    curent_use = reaction_usage_count[reaction_id]

                    # Determine which repeat number to use based on the usage count
                    # 1- reuse_limit: repeat_number = 0 (original)
                    # reuse_limit + 1 - 2x reuse_limit = 1 (first repeat) etc
                    repeat_number = (curent_use -1)//self.reuse_limit

                    # Construct the appropriate reaction_id with suffix
                    if repeat_number == 0:
                        # use original reaction well
                        well_key = (batch_num, reaction_id)

                    else:
                        # Use repeat reaction well
                        well_key = (batch_num, f"{reaction_id}_repeat{repeat_number}")

                    if well_key in reaction_wells:
                        well = reaction_wells[well_key]
                        required_wells.append(well)
                    else:
                        logger.warning(f"Could not find well for {well_key[1]} in batch {batch_num} (use {curent_use}, repeat {repeat_number}, limit {self.reuse_limit})")
                # Create assembly group row
                row = [
                    batch_num,
                    construct_name,
                    ", ".join(required_wells),
                    construct_info['topology']
                ]
                assembly_groups.append(row)

        # Save to file
        assembly_file = f"{output_prefix}_assembly_groups.tsv"
        with open(assembly_file, 'w') as f:
            for row in assembly_groups:
                f.write("\t".join(str(x) for x in row) + "\n")

        logger.info(f"Assembly groups saved to: {assembly_file}")

        # Store for potential machine instruction generation
        self.assembly_groups = assembly_groups

    def generate_epmotion_instructions(self, output_prefix: str, timestamp: str) -> str:
        """Generate instructions for the epMotion liquid handling robot.

        Creates a CSV file containing transfer instructions for primers and templates
        formatted specifically for the epMotion robot. The instructions include:
        - Header information
        - Forward primer transfers
        - Reverse primer transfers
        - Template transfers

        Column order in human_instructions:
        [Batch, Part_name, Constructs, F_Primer, F_Plate, F_Well,
        R_Primer, R_Plate, R_Well, Template, Size, Repeat]

        Args:
            output_prefix: Path prefix for output files (same as human instructions)
            timestamp (str): Timestamp for file naming

        Returns:
            str: Path to the generated instructions file

        Raises:
            ValueError: If assembly instructions haven't been generated
            ValueError: If template information is missing
        """
        if not hasattr(self, 'human_instructions'):
            raise ValueError("No instructions found. Run generate_human_instructions first.")

        # Get unique primer plate barcodes from human instructions
        # Skip header row [1:]
        f_primer_plates = {row[4] for row in self.human_instructions[1:] if row[4] != "N/A"}
        r_primer_plates = {row[7] for row in self.human_instructions[1:] if row[7] != "N/A"}

        # Combine and sort all unique plate barcodes for consistent positions
        all_primer_plates = sorted(f_primer_plates | r_primer_plates)

        # Create epMotion instructions list starting with required headers

        epmotion_instructions = [row[:] for row in self.EPMOTION_HEADER]


        # Process each PCR reaction from human instructions
        # Skip header row [1:]
        for i, row in enumerate(self.human_instructions[1:], 1):
            # Get destination well based on position in list (A1, A2, etc.)
            row_letter = chr(65 + ((i-1) // 12))  # A, B, C, etc.
            col_number = ((i-1) % 12) + 1        # 1, 2, 3, etc.
            pcr_well = f"{row_letter}{col_number}"

            # Add forward primer transfer
            if row[4] != "N/A":  # F_Plate exists
                epmotion_instructions.append([
                    row[4],  # F_Plate (Barcode ID)
                    str(all_primer_plates.index(row[4]) + 1),  # Labware position
                    row[5],  # F_Well (Source)
                    "1",  # Destination plate ID (always 1 for now)
                    pcr_well,  # PCR well position (e.g., A1, B5)
                    "1",  # Volume in µL
                    "TS_50",  # Tool
                    ""  # Name (empty)
                ])

            # Add reverse primer transfer
            if row[7] != "N/A":  # R_Plate exists
                epmotion_instructions.append([
                    row[7],  # R_Plate (Barcode ID)
                    str(all_primer_plates.index(row[7]) + 1),  # Labware position
                    row[8],  # R_Well (Source)
                    "1",  # Destination plate ID
                    pcr_well,  # PCR well position (e.g., A1, B5)
                    "1",  # Volume in µL
                    "TS_50",  # Tool
                    ""  # Name (empty)
                ])

            # Add template transfer if template exists
            if row[9] != "Not found":  # Template exists
                template_plate, template_well = self._get_template_position(row[9])
                if template_plate and template_well:
                    epmotion_instructions.append([
                        template_plate,  # Barcode ID
                        str(len(all_primer_plates) + 1),  # Templates after primers
                        template_well,  # Source well
                        "1",  # Destination plate ID
                        pcr_well,  # PCR well position (e.g., A1, B5)
                        "1",  # Volume in µL
                        "TS_50",  # Tool
                        ""  # Name (empty)
                    ])

        # Save instructions to CSV in same location as human instructions
        output_file = f"{output_prefix}_epmotion_{timestamp}.csv"
        with open(output_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(epmotion_instructions)

        logger.info(f"epMotion instructions saved to: {output_file}")
        return str(output_file)

    def generate_janus_instructions(self, output_prefix: str, timestamp: str) -> str:
        """"Generate instructions for the Janus liquid handling robot

        Creates a CSV file containing transfer instructions for primers and templates
        formatted specifically for the Janus robot. The instructions include:
        - Header information
        - Forward primer transfers
        - Reverse primer transfers
        - Template transfers

        Column order in human_instructions:
        ['construct_id', 'asperate_plate', 'asperate_well', 'destination_plate',
        'destination_well', 'transfer_volume']

        Args:
        output_prefix: Path prefix for output files (same as human instructions)
            timestamp (str): Timestamp for file naming

        Returns:
            str: Path to the generated instructions file

        Raises:
            ValueError: If assembly instructions haven't been generated
            ValueError: If template information is missing
        """
        if not hasattr(self, 'human_instructions'):
            raise ValueError("No instructions found. Run generate_human_instructions first.")

        # Get unique primer plate barcodes from human instructions
        # Skip header row [1:]
        {row[4] for row in self.human_instructions[1:] if row[4] != "N/A"}
        {row[7] for row in self.human_instructions[1:] if row[7] != "N/A"}

        # Combine and sort all unique plate barcodes for consistent positions
        # all_primer_plates = sorted(f_primer_plates | r_primer_plates)

        # Create Janus instructions list starting with required headers
        janus_instructions = [row[:] for row in self.JANUS_HEADER]

        # Process each PCR reaction from human instructions
        # Skip header row [1:]
        for i, row in enumerate(self.human_instructions[1:], 1):
            # Get destination well based on position in list (A1, A2, etc.)
            row_letter = chr(65 + ((i-1) // 12))  # A, B, C, etc.
            col_number = ((i-1) % 12) + 1        # 1, 2, 3, etc.
            pcr_well = f"{row_letter}{col_number}"
            ## to do: link assembly plate # to batch number for v. large assemblies
            # Add forward primer transfer
            if row[4] != "N/A":  # F_Plate exists
                janus_instructions.append([
                    row[2],  # construct id
                    row[4],  # F_Plate (Barcode ID)
                    row[5],  # F_Well (Source)
                    "PCR_plate",  # Destination plate ID (single plate for now)

                    pcr_well,  # PCR well position (e.g., A1, B5)
                    '1'         # Volume in µL

                ])

            # Add reverse primer transfer
            if row[7] != "N/A":  # R_Plate exists
                janus_instructions.append([
                    row[2],  # construct d
                    row[7],  # R_Plate (Barcode ID)
                    row[8],  # R_Well (Source)

                    "PCR_plate",  # Destination plate ID place holder
                        ## to do: link assembly plate # to batch number for v. large assemblies
                    pcr_well,  # PCR well position (e.g., A1, B5)
                    '1',     # Volume in µL
                ])

            # Add template transfer if template exists
            if row[9] != "Not found":  # Template exists
                template_plate, template_well = self._get_template_position(row[9])
                if template_plate and template_well:
                    janus_instructions.append([
                        row[2],         #construct id
                        template_plate, # Barcode ID
                        template_well,  # Source well
                        "PCR_plate",  # Destination plate ID
                        ## to do: link assembly plate # to batch number for v. large assemblies
                        pcr_well,  # PCR well position (e.g., A1, B5)
                        '1',       # volume in µL
                    ])

        # Save instructions in same location as human instructions
        janus_file = f"{output_prefix}_worklist_{timestamp}.csv"
        with open(janus_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(janus_instructions)

        logger.info(f"Worklist saved to: {janus_file}")
        return str(janus_file)


    def _get_template_position(self, template_name: str) -> tuple[Optional[str], Optional[str]]:
        """Find the plate and well position for a given template.

        First checks if the template is a contig/chromosome in the genome mapping file
        to get the genome name, then looks up the genome name in TemPlates.xlsx for position.
        For non-genome templates, directly searches TemPlates.xlsx.
        Searches both public and private directories.

        Args:
            template_name: Name of the template or genome contig to locate

        Returns:
            Tuple containing:
                - Plate name/barcode (or None if not found)
                - Well position (or None if not found)
                Format of well position is standard 96-well notation (e.g., 'A1', 'H12')

        Raises:
            FileNotFoundError: If required mapping files aren't found
        """

        # Construct private template directory path
        private_template_folder = get_templates_path(private=True)

        # Step 1: Check if template_name is a contig and get genome name
        genome_name = None

        for search_dir in [self.template_folder, private_template_folder]:
            if not search_dir.exists():
                continue

            genome_map_file = search_dir / 'genome_well_mapping.tsv'
            if genome_map_file.exists():
                try:
                    genome_df = pd.read_csv(genome_map_file, sep='\t')

                    # Search through each row's contig list
                    for _, row in genome_df.iterrows():
                        contigs = [c.strip() for c in row['Contig_Names'].split(',')]
                        if template_name in contigs:
                            genome_name = row['Genome_Name']
                            break

                    if genome_name:
                        break

                except Exception as e:
                    logger.warning(f"Error reading genome mapping file: {str(e)}")

        # Step 2: Look up the template (or genome name) in TemPlates.xlsx
        # If we found a genome name, search for that; otherwise search for the original template_name
        search_name = genome_name if genome_name else template_name

        for search_dir in [self.template_folder, private_template_folder]:
            if not search_dir.exists():
                continue

            template_excel = search_dir / 'TemPlates.xlsx'
            if not template_excel.exists():
                continue

            try:
                wb = openpyxl.load_workbook(template_excel)

                # Search each sheet (plate)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in range(3, 11):  # Rows 3-10 map to A-H
                        for col in range(2, 14):  # Columns 2-13 map to 1-12
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value == search_name:
                                # Convert Excel row/col to plate coordinates
                                well = f"{chr(row-3+65)}{col-1}"  # Convert 3->A, 4->B, etc.
                                return sheet, well

            except Exception as e:
                logger.warning(f"Error reading template plate map: {str(e)}")

        logger.warning(f"Template {template_name} not found in any mapping files")
        return None, None

    def generate_machine_assembly_instructions(self, output_prefix: str, machine_type: str, timestamp: str) -> str:
        """Generate machine instructions for combining PCR products into final assemblies.

        Args:
            machine_type: Type of liquid handling machine (e.g., 'epmotion')
            output_prefix: Path prefix for output files (same as human instructions)
            timestamp: Timestamp for file naming

        Returns:
            str: Path to the generated instructions file

        Raises:
            ValueError: If assembly groups haven't been generated
            ValueError: If machine type is not supported
        """
        if not hasattr(self, 'assembly_groups'):
            raise ValueError("No assembly groups found. Run generate_assembly_groups first.")

        if machine_type.lower() not in ['epmotion', 'janus']:
            raise ValueError(f"Unsupported machine type: {machine_type}")

        def well_num_to_a1(well_num: int) -> str:
            """Convert well number (1-96) to A1-H12 format assuming horizontal layout."""
            well_num = int(well_num)  # Convert from string if needed
            row = chr(65 + ((well_num - 1) // 12))  # A-H
            col = ((well_num - 1) % 12) + 1         # 1-12
            return f"{row}{col}"

        # Skip header row
        assembly_data = self.assembly_groups[1:]

        # Generate destination wells for assemblies (A1-H12)
        assembly_wells = {}
        for i, assembly in enumerate(assembly_data, 1):
            row_letter = chr(65 + ((i-1) // 12))  # A, B, C, etc.
            col_number = ((i-1) % 12) + 1         # 1, 2, 3, etc.
            assembly_wells[assembly[1]] = f"{row_letter}{col_number}"  # Map construct name to well

        if machine_type.lower() == 'epmotion':
            # Create epMotion format instructions
            instructions = [row[:] for row in self.EPMOTION_HEADER]

            # Process each assembly
            for assembly in assembly_data:
                batch_num = assembly[0]  # Batch number
                construct = assembly[1]  # Construct name
                required_wells = [w.strip().replace('Well ', '') for w in assembly[2].split(',')]  # Get well numbers

                # Add transfers for each required PCR product
                for well_num in required_wells:
                    source_well = well_num_to_a1(well_num)
                    instructions.append([
                        str(batch_num),  # Source plate is batch number
                        str(batch_num),  # Labware position matches batch
                        source_well,     # Source well in A1-H12 format
                        "1",            # Destination plate ID
                        assembly_wells[construct],  # Destination well
                        "2",            # Volume in µL (typical for yeast assembly)
                        "TS_50",        # Tool
                        ""              # Name (empty)
                    ])

            # Save instructions in same location as human instructions
            assembly_file = f"{output_prefix}_assembly_epmotion_{timestamp}.csv"
            with open(assembly_file, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(instructions)

            logger.info(f"epMotion assembly instructions saved to: {assembly_file}")
            return str(assembly_file)

        if machine_type.lower() == 'janus' or machine_type.lower() == 'hamilton':
            # Create janus format instructions
            instructions = [row[:] for row in self.JANUS_HEADER]

            # Process each assembly
            for assembly in assembly_data:
                batch_num = assembly[0]  # Batch number
                construct = assembly[1]  # Construct name
                required_wells = [w.strip().replace('Well ', '') for w in assembly[2].split(',')]  # Get well numbers

                # Add transfers for each required PCR product
                for well_num in required_wells:
                    source_well = well_num_to_a1(well_num)
                    instructions.append([
                        construct,                 # construct id
                        f"plate_{str(batch_num)}", # Source plate is batch number
                        source_well,               # Source well in A1-H12 format
                        "assembly_plate",          # name for the destination plate
                        assembly_wells[construct], # Destination well
                        "2",                       # Volume in µL (typical for yeast assembly)
                    ])

            # Save instructions in same location as human instructions
            assembly_file = f"{output_prefix}_assembly_worklist_{timestamp}.csv"
            with open(assembly_file, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(instructions)

            logger.info(f"Worklist saved to: {assembly_file}")
            return str(assembly_file)

    def save_input_record(self, output_prefix: str) -> None:
        """Save a record of input constructs to file for future reference.

        Args:
            output_prefix: Path prefix for output files
        """
        if not self.selected_constructs:
            logger.warning("No constructs to record")
            return

        # Create table data showing what was selected
        input_data = []
        for name, record in self.selected_constructs.items():
            source = record.annotations.get('source_folder', 'output/')
            source_type = record.annotations.get('source_type', 'unknown')
            topology = record.annotations.get('topology', '').lower()

            # Count components
            if 'PYEAST_component' in [f.type for f in record.features]:
                component_count = sum(1 for f in record.features if f.type == "PYEAST_component")
            else:
                component_count = sum(1 for f in record.features if f.type == "misc_feature")

            input_data.append([
                name,
                source_type,
                source,
                topology,
                f"{len(record)} bp",
                f"{component_count} components"
            ])

        # Save to file
        inputs_file = f"{output_prefix}_inputs.txt"
        with open(inputs_file, 'w') as f:
            f.write("Batch Assembly Input Record\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Selected constructs: {len(self.selected_constructs)}\n\n")
            f.write(tabulate(
                input_data,
                headers=["Name", "Type", "Source Folder", "Topology", "Length", "Components"],
                tablefmt="grid"
            ))

        logger.info(f"Input record saved to: {inputs_file}")
